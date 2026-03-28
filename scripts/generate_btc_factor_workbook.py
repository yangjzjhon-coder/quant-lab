from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CATEGORY_WEIGHTS = {
    "价格趋势结构": 30,
    "量价与订单流": 22,
    "衍生品结构": 22,
    "链上与稳定币流动性": 14,
    "宏观与机构资金": 12,
}


FACTORS = [
    {
        "category": "价格趋势结构",
        "code": "price_vs_200ema_1d",
        "name": "日线200EMA位置",
        "weight": 6,
        "direction": "收盘价在200EMA上方且均线抬头时高分",
        "meaning": "衡量 BTC 是否处在长期多头趋势中，是趋势跟随的总开关之一。",
        "importance": "大级别趋势向上时，中期突破信号的胜率通常更高；跌破时，追多容易变成抄反弹。",
        "scoring": "二元或三段式评分：上方且抬头=1，附近震荡=0.5，下方=0。",
        "frequency": "1D",
        "source": "OKX K线、TradingView、本地 OHLCV",
        "priority": "A",
    },
    {
        "category": "价格趋势结构",
        "code": "ema50_slope_4h",
        "name": "4小时50EMA斜率",
        "weight": 5,
        "direction": "斜率越向上越高分",
        "meaning": "衡量 4 小时中期趋势的上升速度，避免只看均线位置不看斜率。",
        "importance": "同样在均线上方，斜率抬升的行情更容易演化为持续趋势；走平时更容易来回震荡。",
        "scoring": "连续评分：斜率按 ATR 或价格归一化后映射到 0 到 1。",
        "frequency": "4H",
        "source": "OKX K线、本地 OHLCV",
        "priority": "A",
    },
    {
        "category": "价格趋势结构",
        "code": "multi_tf_alignment",
        "name": "多周期趋势一致性",
        "weight": 6,
        "direction": "1H/4H/1D 同向时高分",
        "meaning": "同时比较短、中、长三个周期的趋势方向是否一致。",
        "importance": "多周期同向时，趋势延续概率更高；周期打架时，假突破和回撤频率更高。",
        "scoring": "1D、4H、1H 三个周期每个给 0 或 1，再按平均值映射。",
        "frequency": "1H / 4H / 1D",
        "source": "OKX K线、本地 OHLCV",
        "priority": "A",
    },
    {
        "category": "价格趋势结构",
        "code": "donchian_breakout_strength",
        "name": "唐奇安突破强度",
        "weight": 5,
        "direction": "向上突破且距离适中时高分",
        "meaning": "衡量价格是否有效突破过去一段时间高点，以及突破是否有延续空间。",
        "importance": "趋势跟随本质上是追随新高或新低，突破强度决定了入场是否顺着主趋势。",
        "scoring": "结合突破距离、回踩位置和 ATR 归一化后评分。",
        "frequency": "4H / 1D",
        "source": "OKX K线、本地 OHLCV",
        "priority": "A",
    },
    {
        "category": "价格趋势结构",
        "code": "adx_4h_strength",
        "name": "4小时ADX趋势强度",
        "weight": 4,
        "direction": "ADX 越高越高分，但过热区可降分",
        "meaning": "衡量趋势是否真的在扩展，而不是均线刚刚金叉后的假动作。",
        "importance": "ADX 能过滤大量横盘假信号，是趋势系统很常见的强度过滤器。",
        "scoring": "分段评分：低于阈值 0，中等强度 0.5 到 0.8，强趋势 1。",
        "frequency": "4H",
        "source": "OKX K线、本地 OHLCV",
        "priority": "A",
    },
    {
        "category": "价格趋势结构",
        "code": "atr_expansion_state",
        "name": "ATR波动扩张状态",
        "weight": 4,
        "direction": "突破后温和扩张时高分",
        "meaning": "衡量波动率是否从压缩转向扩张，确认市场开始给趋势腾出空间。",
        "importance": "趋势经常从低波动转向中高波动；没有波动扩张，追涨往往走不远。",
        "scoring": "当前 ATR 相对历史中位数或均值的比值映射到 0 到 1。",
        "frequency": "4H",
        "source": "OKX K线、本地 OHLCV",
        "priority": "B",
    },
    {
        "category": "量价与订单流",
        "code": "breakout_volume_ratio",
        "name": "突破量能放大比",
        "weight": 6,
        "direction": "放量突破时高分",
        "meaning": "衡量突破对应的成交量是否显著高于近期均值。",
        "importance": "没有量能配合的突破容易失败，放量说明市场愿意在更高价继续成交。",
        "scoring": "当前成交量 / 滚动均量后做阈值或连续映射。",
        "frequency": "1H / 4H",
        "source": "OKX K线、逐笔成交聚合",
        "priority": "A",
    },
    {
        "category": "量价与订单流",
        "code": "obv_slope",
        "name": "OBV趋势斜率",
        "weight": 4,
        "direction": "OBV 抬升越明显越高分",
        "meaning": "用成交量净累积来验证价格上涨是否有持续买盘跟随。",
        "importance": "价格新高但 OBV 不创新高，往往说明趋势质量下降。",
        "scoring": "OBV 均线斜率或 OBV 新高确认做 0 到 1 评分。",
        "frequency": "4H",
        "source": "OKX K线、本地 OHLCV",
        "priority": "B",
    },
    {
        "category": "量价与订单流",
        "code": "cvd_delta",
        "name": "CVD累积成交量差",
        "weight": 5,
        "direction": "主动买入占优时高分",
        "meaning": "衡量主动买盘与主动卖盘的净差额，判断到底是谁在推价格。",
        "importance": "趋势启动期如果价格上涨同时 CVD 上行，说明上冲不是被动抬价而是真买盘推动。",
        "scoring": "短中周期 CVD 斜率、创新高确认或背离情况映射到 0 到 1。",
        "frequency": "5m / 15m / 1H",
        "source": "交易所逐笔成交、WebSocket 成交流",
        "priority": "A",
    },
    {
        "category": "量价与订单流",
        "code": "anchored_vwap_position",
        "name": "锚定VWAP位置",
        "weight": 4,
        "direction": "价格在关键锚定 VWAP 上方时高分",
        "meaning": "以最近重要低点、突破点或周期开盘为锚点，衡量市场平均持仓成本位置。",
        "importance": "价格站稳关键锚定 VWAP，说明新增持仓群体整体处于盈利状态，趋势更稳。",
        "scoring": "价格相对锚定 VWAP 的距离和回踩有效性综合评分。",
        "frequency": "1H / 4H",
        "source": "K线数据、自定义锚点",
        "priority": "B",
    },
    {
        "category": "量价与订单流",
        "code": "large_trade_net_aggression",
        "name": "大单净主动买入占比",
        "weight": 3,
        "direction": "大单主动买入越多越高分",
        "meaning": "统计大额成交中主动买单和主动卖单的差异，观察主力是否在进场。",
        "importance": "趋势往往需要大资金推动；如果小单拉升而大单砸盘，趋势质量通常较差。",
        "scoring": "设定大单阈值后，净主动买入金额占比映射到 0 到 1。",
        "frequency": "1m / 5m / 15m",
        "source": "逐笔成交、盘口成交明细",
        "priority": "B",
    },
    {
        "category": "衍生品结构",
        "code": "oi_price_confirmation",
        "name": "未平仓量与价格同向确认",
        "weight": 6,
        "direction": "价格涨且 OI 同步增时高分",
        "meaning": "观察上涨是否伴随新增仓位进入，而不是单纯空头回补。",
        "importance": "趋势跟随更喜欢新增资金推动的趋势，而不是短期轧空后的脉冲行情。",
        "scoring": "价格变化与 OI 变化的同向性做连续评分，背离时降分。",
        "frequency": "1H / 4H",
        "source": "OKX、Coinglass、交易所 OI 接口",
        "priority": "A",
    },
    {
        "category": "衍生品结构",
        "code": "basis_term_structure",
        "name": "基差与期限结构",
        "weight": 5,
        "direction": "温和正基差时高分，极端升水降分",
        "meaning": "衡量现货与期货价格差，以及不同期限合约之间的结构是否健康。",
        "importance": "适度升水说明市场乐观，过高升水往往意味着拥挤交易和回撤风险上升。",
        "scoring": "基差落在健康区间给高分，过热或倒挂给低分。",
        "frequency": "4H / 1D",
        "source": "OKX 交割合约、第三方衍生品数据",
        "priority": "A",
    },
    {
        "category": "衍生品结构",
        "code": "funding_regime",
        "name": "资金费率健康度",
        "weight": 4,
        "direction": "中性偏多时高分，极端正费率降分",
        "meaning": "观察永续合约多头是否过于拥挤，避免在极端乐观时高位接盘。",
        "importance": "资金费率过高通常意味着趋势已被透支，适合控仓而不是激进追价。",
        "scoring": "健康区间 1，极端正费率或极端负费率逐步降到 0。",
        "frequency": "8H / 4H",
        "source": "OKX 永续资金费率",
        "priority": "A",
    },
    {
        "category": "衍生品结构",
        "code": "liquidation_pressure_map",
        "name": "清算密集区压力",
        "weight": 4,
        "direction": "上方有可推升空间且下方清算压力有限时高分",
        "meaning": "观察杠杆集中区和潜在清算带，判断价格是否容易被挤压加速。",
        "importance": "趋势延续时，清算带常常成为加速器；反过来，贴近下方清算带时风险更大。",
        "scoring": "按上下方清算热区相对位置和密度做 0 到 1 评分。",
        "frequency": "1H / 4H",
        "source": "Coinglass、订单簿热力图、清算地图",
        "priority": "B",
    },
    {
        "category": "衍生品结构",
        "code": "options_skew_iv_term",
        "name": "期权偏度与隐波期限结构",
        "weight": 3,
        "direction": "风险偏好改善时高分",
        "meaning": "观察看涨看跌偏度与近远月隐波结构，判断机构是在对冲还是追涨。",
        "importance": "如果现货上涨但期权市场明显偏向保护性看跌，趋势持续性通常要打折。",
        "scoring": "偏度和期限结构落在健康区间时给高分。",
        "frequency": "1D",
        "source": "Deribit、Laevitas、Greeks 数据平台",
        "priority": "C",
    },
    {
        "category": "链上与稳定币流动性",
        "code": "btc_exchange_netflow",
        "name": "BTC交易所净流入流出",
        "weight": 4,
        "direction": "净流出时高分",
        "meaning": "观察 BTC 是被转入交易所准备卖出，还是被提出交易所进入持有状态。",
        "importance": "大级别净流出通常意味着卖压减轻，净流入则可能意味着上方抛压增加。",
        "scoring": "滚动净流量相对历史分位数映射到 0 到 1。",
        "frequency": "1D",
        "source": "Glassnode、CryptoQuant、交易所钱包监控",
        "priority": "B",
    },
    {
        "category": "链上与稳定币流动性",
        "code": "stablecoin_exchange_inflow",
        "name": "稳定币交易所净流入",
        "weight": 4,
        "direction": "净流入增加时高分",
        "meaning": "衡量可用于买入 BTC 的场内购买力是否在增强。",
        "importance": "稳定币净流入增加通常意味着场内火力增强，是推动趋势的重要燃料。",
        "scoring": "按净流入规模和持续性做 0 到 1 评分。",
        "frequency": "1D",
        "source": "Glassnode、CryptoQuant、链上稳定币监控",
        "priority": "B",
    },
    {
        "category": "链上与稳定币流动性",
        "code": "sopr_regime",
        "name": "SOPR盈利兑现状态",
        "weight": 3,
        "direction": "重回 1 上方并稳定时高分",
        "meaning": "衡量链上筹码是否整体回到盈利状态，以及获利了结压力是否可控。",
        "importance": "SOPR 站稳 1 上方通常代表市场愿意在盈利状态下继续持有，是牛市趋势的重要特征。",
        "scoring": "高于或低于关键阈值做三段式评分。",
        "frequency": "1D",
        "source": "Glassnode、链上因子平台",
        "priority": "C",
    },
    {
        "category": "链上与稳定币流动性",
        "code": "miner_whale_sell_pressure",
        "name": "矿工与鲸鱼卖压",
        "weight": 3,
        "direction": "卖压下降时高分",
        "meaning": "跟踪矿工和大额地址向交易所转币的行为，评估潜在抛售供给。",
        "importance": "趋势行情里如果鲸鱼和矿工持续往交易所转币，往往会压制上行持续性。",
        "scoring": "卖压越小越接近 1，集中抛压时接近 0。",
        "frequency": "1D",
        "source": "链上地址监控、Glassnode、CryptoQuant",
        "priority": "C",
    },
    {
        "category": "宏观与机构资金",
        "code": "spot_etf_netflow",
        "name": "现货ETF净流入",
        "weight": 4,
        "direction": "持续净流入时高分",
        "meaning": "跟踪美国现货 BTC ETF 的净申购和净赎回，观察机构资金是否在持续入场。",
        "importance": "ETF 资金是近年推动 BTC 趋势的重要新变量，持续净流入会明显改善上行质量。",
        "scoring": "按近 5 到 20 日净流入强弱分段或连续评分。",
        "frequency": "1D",
        "source": "Farside、基金官网、Bloomberg 汇总",
        "priority": "A",
    },
    {
        "category": "宏观与机构资金",
        "code": "dxy_regime",
        "name": "美元指数DXY趋势",
        "weight": 3,
        "direction": "DXY 走弱时高分",
        "meaning": "衡量美元是否持续走强，因为强美元通常会压制风险资产表现。",
        "importance": "BTC 和美元流动性环境高度相关，DXY 走弱常常对风险资产更友好。",
        "scoring": "DXY 相对自身均线和斜率做反向评分。",
        "frequency": "1D",
        "source": "TradingView、FRED、宏观行情接口",
        "priority": "B",
    },
    {
        "category": "宏观与机构资金",
        "code": "nasdaq_risk_appetite",
        "name": "纳指风险偏好联动",
        "weight": 3,
        "direction": "纳指偏强时高分",
        "meaning": "观察 BTC 是否处在与成长股风险偏好共振的环境里。",
        "importance": "不少阶段里 BTC 会和高贝塔科技资产同涨同跌，风险偏好共振会强化趋势持续性。",
        "scoring": "纳指趋势、波动率和相关性组合映射到 0 到 1。",
        "frequency": "1D",
        "source": "纳指指数、TradingView、券商行情接口",
        "priority": "B",
    },
    {
        "category": "宏观与机构资金",
        "code": "real_yield_pressure",
        "name": "实际利率与美债压力",
        "weight": 2,
        "direction": "实际利率回落时高分",
        "meaning": "关注美国 10 年期实际利率或名义利率是否对高风险资产形成压制。",
        "importance": "利率上行会提高持有无现金流资产的机会成本，对 BTC 估值扩张不友好。",
        "scoring": "按利率趋势和变化幅度做反向评分。",
        "frequency": "1D",
        "source": "FRED、宏观数据库、券商行情接口",
        "priority": "C",
    },
]


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_header(row) -> None:
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def set_widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def build_factor_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("因子总表")
    headers = [
        "序号",
        "因子类别",
        "类别权重",
        "因子代码",
        "因子名称",
        "因子权重",
        "主要方向",
        "中文含义",
        "为什么重要",
        "建议评分方式",
        "建议频率",
        "数据来源示例",
        "测试优先级",
    ]
    ws.append(headers)
    style_header(ws[1])

    for index, factor in enumerate(FACTORS, start=1):
        ws.append(
            [
                index,
                factor["category"],
                CATEGORY_WEIGHTS[factor["category"]],
                factor["code"],
                factor["name"],
                factor["weight"],
                factor["direction"],
                factor["meaning"],
                factor["importance"],
                factor["scoring"],
                factor["frequency"],
                factor["source"],
                factor["priority"],
            ]
        )
        name_cell = ws.cell(row=index + 1, column=5)
        name_cell.comment = Comment(
            text=f"中文注释：{factor['meaning']}\n趋势用途：{factor['importance']}",
            author="Codex",
        )

    total_row = len(FACTORS) + 2
    ws.cell(row=total_row, column=4, value="总权重")
    ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{len(FACTORS)+1})")
    for cell in ws[total_row]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{len(FACTORS)+1}"
    style_body(ws)
    set_widths(
        ws,
        {
            "A": 8,
            "B": 16,
            "C": 10,
            "D": 28,
            "E": 18,
            "F": 10,
            "G": 24,
            "H": 28,
            "I": 32,
            "J": 30,
            "K": 14,
            "L": 24,
            "M": 10,
        },
    )


def build_category_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("类别权重")
    ws.append(["类别", "类别权重", "包含因子数", "说明"])
    style_header(ws[1])

    descriptions = {
        "价格趋势结构": "决定是否允许开仓，是趋势跟随的核心层。",
        "量价与订单流": "确认突破质量和主动买盘强度，负责提升胜率。",
        "衍生品结构": "识别杠杆拥挤、资金拥挤与新增仓位推动情况。",
        "链上与稳定币流动性": "判断场内外真实购买力和潜在抛压供给。",
        "宏观与机构资金": "识别 BTC 所处的大环境和机构资金边际变化。",
    }

    row_index = 2
    for category, weight in CATEGORY_WEIGHTS.items():
        count = sum(1 for factor in FACTORS if factor["category"] == category)
        ws.append([category, weight, count, descriptions[category]])
        row_index += 1

    ws.append(["总计", "=SUM(B2:B6)", f"=SUM(C2:C6)", "总分数固定为 100 分"])
    for cell in ws[row_index]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    style_body(ws)
    set_widths(ws, {"A": 20, "B": 12, "C": 12, "D": 42})


def build_strategy_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("趋势整合建议")
    ws.append(["模块", "内容"])
    style_header(ws[1])

    rows = [
        ("策略目标", "先做 BTC 多头趋势跟随，空头后续单独建模，不和多头共用一套镜像规则。"),
        ("总分公式", "总分 = 所有单因子评分(0到1) × 因子权重 之和，满分 100 分。"),
        ("一级门槛", "日线200EMA位置、多周期一致性、唐奇安突破强度 这三项合计建议至少达到 12 分，否则不追趋势。"),
        ("二级确认", "量能放大、CVD、OI与价格同向、ETF净流入 这四项至少有两项明显偏多，才允许正常仓位开仓。"),
        ("风险约束", "资金费率过热、基差过热、下方清算带过近时，即使总分高也要降杠杆、减仓或等待回踩。"),
        ("建议分层", "70到100 分：强趋势，可正常仓位跟随；55到69 分：轻仓试多或等确认；40到54 分：观察，不开新仓；40 分以下：禁止开新仓。"),
        ("持仓管理", "开仓后优先看价格趋势结构和量价确认是否持续；若总分跌破 50 且趋势核心项明显恶化，可主动减仓。"),
        ("测试顺序", "先测价格趋势结构 + 量能 + OI 三层，再逐步加入链上、ETF、宏观，不建议一开始全量堆因子。"),
        ("因子冲突处理", "趋势核心层权重大于确认层，确认层大于环境层；核心层转空时，其它利多因子只允许降空仓，不允许硬做多。"),
        ("建议研发节奏", "第一阶段只接 OKX 可直接拿到的数据和 ETF/DXY 外部数据；第二阶段再接链上与期权数据。"),
    ]

    for row in rows:
        ws.append(list(row))

    style_body(ws)
    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 18, "B": 92})


def build_template_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("打分模板")
    headers = [
        "序号",
        "因子类别",
        "因子名称",
        "因子权重",
        "当前评分(0到1)",
        "加权得分",
        "备注",
    ]
    ws.append(headers)
    style_header(ws[1])

    for index, factor in enumerate(FACTORS, start=2):
        ws.append(
            [
                index - 1,
                factor["category"],
                factor["name"],
                factor["weight"],
                "",
                f"=D{index}*E{index}",
                "",
            ]
        )

    total_row = len(FACTORS) + 2
    ws.cell(row=total_row, column=3, value="总分")
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})")
    ws.cell(row=total_row, column=7, value="目标满分 100")
    for cell in ws[total_row]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    style_body(ws)
    set_widths(ws, {"A": 8, "B": 16, "C": 18, "D": 10, "E": 14, "F": 12, "G": 24})


def add_cover_sheet(workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "说明"
    ws.append(["BTC 趋势跟随因子测试框架"])
    ws.append(["版本日期", "2026-03-26"])
    ws.append(["用途", "用于先整理因子池、权重和打分框架，不代表已经完成回测验证。"])
    ws.append(["总分规则", "所有因子权重合计 100 分，单因子先按 0 到 1 打分，再乘权重。"])
    ws.append(["当前定位", "优先服务于 BTC 多头趋势跟随；空头建议单独建模。"])
    ws.append(["备注", "因子名称单元格已加中文批注，便于在 Excel 中悬停查看。"])

    ws["A1"].font = Font(bold=True, size=14, color="1F1F1F")
    ws["A1"].fill = SUBHEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:B1")

    style_body(ws)
    set_widths(ws, {"A": 18, "B": 90})


def add_table_styling(workbook: Workbook) -> None:
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.border = THIN_BORDER
        for row_index in range(2, ws.max_row + 1):
            if row_index % 2 == 0 and ws.title in {"因子总表", "打分模板"}:
                for cell in ws[row_index]:
                    if cell.fill == HEADER_FILL:
                        continue
                    cell.fill = PatternFill("solid", fgColor="F8FBFF")


def validate_weights() -> None:
    total_weight = sum(item["weight"] for item in FACTORS)
    if total_weight != 100:
        raise ValueError(f"Factor weights must sum to 100, got {total_weight}.")
    for category, category_weight in CATEGORY_WEIGHTS.items():
        factor_sum = sum(item["weight"] for item in FACTORS if item["category"] == category)
        if factor_sum != category_weight:
            raise ValueError(
                f"Category {category} weight mismatch: category={category_weight}, factors={factor_sum}."
            )


def main() -> None:
    validate_weights()

    project_root = Path(__file__).resolve().parent.parent
    output_dir = project_root / "data" / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "BTC趋势跟随因子框架_2026-03-26.xlsx"

    workbook = Workbook()
    add_cover_sheet(workbook)
    build_factor_sheet(workbook)
    build_category_sheet(workbook)
    build_strategy_sheet(workbook)
    build_template_sheet(workbook)
    add_table_styling(workbook)

    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = True

    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
